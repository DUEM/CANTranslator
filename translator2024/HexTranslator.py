import os
import struct
import sys, getopt
from openpyxl import load_workbook, Workbook
from re import sub, compile
from io import open

configFile = ''
inputFile = ''
outputFile = ''
try:
    opts, args = getopt.getopt(sys.argv[1:],"c:i:o:",["configTable=", "inputBinary=", "outputFile="]) #from https://www.tutorialspoint.com/python/python_command_line_arguments.htm.
except:
    print('Translator.py -i <input data> -o <output file> -c <config table file>')
    sys.exit(2)
for opt, arg in opts:
    #print(arg)
    if opt in ("-c", "--configTable"):
        configFile = arg
    elif opt in ("-h", "--headerStream"):
        headerFile = arg
print("Config file: " + configFile)
print("Input file: " + inputFile)
print("Output file: " + outputFile)

#load configFile
wb = load_workbook(configFile, read_only=True)
sheet = wb.active

if(inputFile != ''):
    print("No input location for hex binary specified")
    sys.exit(3)
if(outputFile == ''):
    print("No output location for translated hex dump specified")
    sys.exit(4)
print("Translating HEX dump")
    
hexOutBook = Workbook()
hexOut = hexOutBook.active
hexOut.title = inputFile.split('\.')[0]

with open(inputFile) as inputStream:
    rowCounter = 2
    for line in inputStream.readlines()[5:]:
        hexValues = line.split()

        #get message format
        msgId = '0x' + ''.join(e for e in hexValues[1:3])
        print(hexValues[1:3])
        print(msgId)
        configRow = -1
        for i in range(2, sheet.max_row):
            print(sheet["E{}".format(i)].value)
            if(sheet["E{}".format(i)].value == msgId): #assuming all IDs are valid. No need to check if row was not found
                configRow = i
                break
        
        #write source and item name
        hexOut["A{}".format(rowCounter)] = sheet["B{}".format(configRow)].value #Item name
        hexOut["B{}".format(rowCounter)] = sheet["C{}".format(configRow)].value #Source of message
        
        #write values
        match sheet["R{}".format(configRow)].value:
            case '2*float32':
                hexOut["C{}".format(rowCounter)] = struct.unpack('<f', bytes.fromhex(''.join(str(element) for element in line[4:8])))[0] #where < means little endian
                hexOut["D{}".format(rowCounter)] = struct.unpack('<f', bytes.fromhex(''.join(str(element) for element in line[8:12])))[0]
            case '3*u_int16': #using H for unsigned short (equal to u_int16)
                hexOut["C{}".format(rowCounter)] = struct.unpack('<H', bytes.fromhex(''.join(str(element) for element in line[4:6])))[0]
                hexOut["D{}".format(rowCounter)] = struct.unpack('<H', bytes.fromhex(''.join(str(element) for element in line[6:8])))[0]
                hexOut["E{}".format(rowCounter)] = struct.unpack('<H', bytes.fromhex(''.join(str(element) for element in line[8:10])))[0]
            case 'u_int32 & char[4]':
                hexOut["C{}".format(rowCounter)] = struct.unpack('<I', bytes.fromhex(''.join(str(element) for element in line[4:8])))[0]
                hexOut["D{}".format(rowCounter)] = ''.join(struct.unpack('<b', b)[0] for b in line[8:12])
            case _:
                print("Undefined data format: " + str(sheet["R{}".format(configRow)].value))
                hexOutBook.save(filename = outputFile + ".xlsx") #save what was successfully translated
                sys.exit(6)

hexOutBook.save(filename = outputFile + ".xlsx")
