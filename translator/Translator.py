import os
#import struct
import sys, getopt
from openpyxl import load_workbook, Workbook
from re import sub, compile
from io import open

#handle command line arguments
#inputfile = ''
#outputfile = ''
configFile = ''
headerFile = ''
try:
    opts, args = getopt.getopt(sys.argv[1:],"c:h:",["configTable=", "headerStream="]) #from https://www.tutorialspoint.com/python/python_command_line_arguments.htm.
except:
    print('Translator.py -i <input data> -o <output file> -c <config table file> -h <header file>')
    sys.exit(2)
for opt, arg in opts:
    #print(arg)
    if opt in ("-c", "--configTable"):
        configFile = arg
    elif opt in ("-h", "--headerStream"):
        headerFile = arg
print("Config file: " + configFile)
print("Header file: " + headerFile)

#load configFile
wb = load_workbook(configFile, read_only=True)
sheet = wb.active

#if header file is specified, then generate a header file (and a cpp file) that enables the message formats specified in the config file to be used in C++ code
if(headerFile != ''):
    def toCamelCase(s): #https://www.w3resource.com/python-exercises/string/python-data-type-string-exercise-96.php
        s1 = sub(r'[^a-zA-Z0-9]+', ' ', s).title().replace(" ", "")
        return s1
    
    if(os.path.isfile(headerFile + '.hpp')):
       os.remove(headerFile + '.hpp')
    if(os.path.isfile(headerFile + '.cpp')):
        os.remove(headerFile + '.cpp')
    headerStream = open(headerFile + '.hpp', 'w')
    implStream = open(headerFile + ".cpp", 'w')
    
    #headerStream.write("#include \"structDefinitions.hpp\"\n#include \"mcp2515.h\"\n")
    headerStream.write("#ifndef " + os.path.basename(headerFile) + "_INCLUDE_GUARD\n")
    headerStream.write("#define " + os.path.basename(headerFile) + "_INCLUDE_GUARD\n")
    headerStream.write("#include \"CANHelper.hpp\"\n")
    #headerStream.write("namespace CANHelper\n{\n\tvoid DispatchMsg(can_frame msg);\n}\n")
    headerStream.write("namespace CANHelper::Messages\n{\n")
    implStream.write('#include \"' + os.path.basename(headerFile) + '.hpp\"\nnamespace CANHelper\n{\n\tvoid CanMsgHandler::DispatchMsg(can_frame msg)\n\t{\n\t\tswitch(msg.can_id)\n\t\t{\n')

    #create header file with class declarations for each record in the config
    rowNumber = 1
    while not (str(sheet.cell(rowNumber, 1).value) == "END"):
        row = [cell.value for cell in sheet[rowNumber][0:18]] #columns 0 to 17 contains table. Anything beyond 17 are just notes and not relevant
        rowNumber = rowNumber + 1
        print(row) # list of cell values of this row
        #print(row[1] + "->" + sub(r"(|-)+.", " ", str(row[1])))
        #print(row[1] + '->' + toCamelCase(row[1]))
        #print(row[8] + "->" + toCamelCase(row[8]))
        headerStream.write("#ifdef USE_MSG_" + toCamelCase(row[2]) + "_" + toCamelCase(row[1]) + "\n")
        headerStream.write("#define CAN_ID_" + toCamelCase(row[2]) + "_" + toCamelCase(row[1]) + " " + str(row[4]) + "\n")
        headerStream.write("#define CAN_DLC_" + toCamelCase(row[2]) + "_" + toCamelCase(row[1]) + " " + str(row[6]) + "\n")
        headerStream.write("\tnamespace " + toCamelCase(row[2]) + "\n\t{\n")
        headerStream.write("\t\tstruct _" + toCamelCase(row[1]) + " : public Messages::CANMsg {\n")
        #headerStream.write("\t\t\tMsgMetadata metadata;\n")
        headerStream.write("\t\t\tstruct canData\n\t\t\t{\n")

        implStream.write("#ifdef USE_MSG_" + toCamelCase(row[2]) + "_" + toCamelCase(row[1]) + "\n")
        implStream.write('\t\tcase ' + row[4] + ':\n')
        implStream.write("\t\t\tMessages::" + "processMessage((Messages::" + toCamelCase(row[2]) + "::_" + toCamelCase(row[1]) + "&)msg);\n")
        implStream.write("\t\t\tbreak;\n")

        #print(row[17])
        #check validity of datatype. TODO: check that datatype matches the DLC.
        datatype = row[17].replace(' ', '')
        if not compile("^([1-8]\*)?([a-z]|[1-9])(,([1-8]\*)?([a-z]|[1-9]))*").match(datatype):
            print("Invalid datatype", end='')
            print(datatype)
            sys.exit(4)
        
        byteLabelIndex = 7 #data description starts at index 7
        for type in datatype.split(','):
            temp = type.split('*')
            noOfRepeats = 1
            typeLabel = temp[0]
            if len(temp) == 2:
                noOfRepeats = int(temp[0])
                typeLabel = temp[1]
            for i in range(0, noOfRepeats):
                while row[byteLabelIndex] == 'IN USE' or row[byteLabelIndex] == '-':
                    byteLabelIndex = byteLabelIndex + 1
                headerStream.write("\t\t\t\t" + typeLabel + " " + toCamelCase(row[byteLabelIndex]) + ";\n")
                byteLabelIndex = byteLabelIndex + 1

        headerStream.write("\t\t\t} __attribute__((aligned(8)));\n")
        headerStream.write("\t\t\tcanData data;\n")
        headerStream.write("\t\t\t_" + toCamelCase(row[1]) + "() : CANMsg(" + "CAN_ID_" + toCamelCase(row[2]) + "_" + toCamelCase(row[1]) + ", " + "CAN_DLC_" + toCamelCase(row[2]) + "_" + toCamelCase(row[1]) + ") { }\n")
        headerStream.write("\t\t};\n\t}\n")
        headerStream.write("\tvoid processMessage(" + toCamelCase(row[2]) + '::_' + toCamelCase(row[1]) + "& msg);\n")
        headerStream.write("#endif\n")
        implStream.write("#endif\n")

    #process all message declaration
    headerStream.write("#ifdef PROCESS_ALL_MSG\n")
    headerStream.write("\tvoid processAll(CANMsg& msg);\n")
    headerStream.write("#endif\n")

    headerStream.write("}\n")
    headerStream.write("#endif")
    headerStream.close()

    implStream.write("\t\t}\n")
    implStream.write("#ifdef PROCESS_ALL_MSG\n")
    implStream.write("\t\tprocessAll((Messages::CANMsg&) msg);\n")
    implStream.write("#endif\n")
    implStream.write("\t}\n}\n")
    implStream.close()
    print("c++ files generated file generated")

    #match typeLabel:
    #            case 'float32':
    #                headerStream.write("float " + toCamelCase(row[byteLabelIndex]))
    #                byteLabelIndex = byteLabelIndex + 4
    #            case _:
    #                print("Unrecognised type ", end='')
    #                print(typeLabel)
    #                sys.exit(3)

    # match row[17]: #col R is index 17
        #     case '2*float32':
        #         headerStream.write("\t\t\t\tfloat " + toCamelCase(row[7]) + ";\n") #col H is index 7
        #         headerStream.write("\t\t\t\tfloat " + toCamelCase(row[11]) + ";\n") #col L is index 11
        #     case '3*u_int16':
        #         headerStream.write("\t\t\t\tuint16_t " + toCamelCase(row[7]) + ";\n")
        #         headerStream.write("\t\t\t\tuint16_t " + toCamelCase(row[9]) + ";\n") #col J is index 9
        #         headerStream.write("\t\t\t\tuint16_t " + toCamelCase(row[11]) + ";\n")
        #     case 'float32, int32':
        #         headerStream.write("\t\t\t\tfloat " + toCamelCase(row[7]) + ";\n")
        #         headerStream.write("\t\t\t\tint " + toCamelCase(row[11]) + ";\n")
        #     case 'u_int32 & char[4]':
        #         headerStream.write("\t\t\t\tuint32_t " + toCamelCase(row[7]) + ";\n")
        #         headerStream.write("\t\t\t\tchar* " + toCamelCase(row[11]) + ";\n")
        #     case 'int16 & 2*u_int16 & 2*u_int8':
        #         headerStream.write("\t\t\t\tint16_t " + toCamelCase(row[7]) + ";\n")
        #         headerStream.write("\t\t\t\tuint16_t " + toCamelCase(row[9]) + ";\n")
        #         headerStream.write("\t\t\t\tuint16_t " + toCamelCase(row[11]) + ";\n")
        #         headerStream.write("\t\t\t\tuint8_t " + toCamelCase(row[13]) + ";\n") #col N is index 13
        #         headerStream.write("\t\t\t\tuint8_t " + toCamelCase(row[14]) + ";\n") #col O is index 14
        #     case '6*int8':
        #         headerStream.write("\t\t\t\t//WIP\n")
        #     case '5*u_int8':
        #         headerStream.write("\t\t\t\tuint8_t " + toCamelCase(row[7]) + ";\n")
        #         headerStream.write("\t\t\t\tuint8_t " + toCamelCase(row[8]) + ";\n")
        #         headerStream.write("\t\t\t\tuint8_t " + toCamelCase(row[9]) + ";\n")
        #         headerStream.write("\t\t\t\tuint8_t " + toCamelCase(row[10]) + ";\n") #col N is index 13
        #         headerStream.write("\t\t\t\tuint8_t " + toCamelCase(row[11]) + ";\n") #col O is index 14
        #     case _:
        #         print("Unrecognised data structure ", end='')
        #         print(row[17])
        #         headerStream.close()
        #         sys.exit(3)
        
        #headerStream.write("\t\t\tinline can_id id() { return CAN_ID_" + toCamelCase(row[2]) + "_" + toCamelCase(row[1]) + "; }\n")
        #headerStream.write("\t\t\tinline __u8 dlc() { return CAN_DLC_" + toCamelCase(row[2]) + "_" + toCamelCase(row[1]) + "; }\n")

#if input file is specified, then translate its contents to something that can be read and processed. It should be a hex dump of CAN messages
# if(inputfile != ''):
#     if(outputfile == ''):
#         print("No ouput location for translated hex dump specified")
#         sys.exit(4)
#     print("Translating HEX dump")
    
#     hexOutBook = Workbook()
#     hexOut = hexOutBook.active
#     hexOut.title = inputfile.split('\.')[0]

#     with open(inputfile) as inputStream:
#         rowCounter = 2
#         for line in inputStream.readlines()[5:]:
#             hexValues = line.split()

#             #get message format
#             msgId = '0x' + ''.join(e for e in hexValues[1:3])
#             print(hexValues[1:3])
#             print(msgId)
#             configRow = -1
#             for i in range(2, sheet.max_row):
#                 print(sheet["E{}".format(i)].value)
#                 if(sheet["E{}".format(i)].value == msgId): #assuming all IDs are valid. No need to check if row was not found
#                     configRow = i
#                     break
            
#             #write source and item name
#             hexOut["A{}".format(rowCounter)] = sheet["B{}".format(configRow)].value #Item name
#             hexOut["B{}".format(rowCounter)] = sheet["C{}".format(configRow)].value #Source of message
            
#             #write values
#             match sheet["R{}".format(configRow)].value:
#                 case '2*float32':
#                     hexOut["C{}".format(rowCounter)] = struct.unpack('<f', bytes.fromhex(''.join(str(element) for element in line[4:8])))[0] #where < means little endian
#                     hexOut["D{}".format(rowCounter)] = struct.unpack('<f', bytes.fromhex(''.join(str(element) for element in line[8:12])))[0]
#                 case '3*u_int16': #using H for unsigned short (equal to u_int16)
#                     hexOut["C{}".format(rowCounter)] = struct.unpack('<H', bytes.fromhex(''.join(str(element) for element in line[4:6])))[0]
#                     hexOut["D{}".format(rowCounter)] = struct.unpack('<H', bytes.fromhex(''.join(str(element) for element in line[6:8])))[0]
#                     hexOut["E{}".format(rowCounter)] = struct.unpack('<H', bytes.fromhex(''.join(str(element) for element in line[8:10])))[0]
#                 case 'u_int32 & char[4]':
#                     hexOut["C{}".format(rowCounter)] = struct.unpack('<I', bytes.fromhex(''.join(str(element) for element in line[4:8])))[0]
#                     hexOut["D{}".format(rowCounter)] = ''.join(struct.unpack('<b', b)[0] for b in line[8:12])
#                 case _:
#                     print("Undefined data format: " + str(sheet["R{}".format(configRow)].value))
#                     hexOutBook.save(filename = outputfile + ".xlsx") #save what was successfully translated
#                     sys.exit(6)

#     hexOutBook.save(filename = outputfile + ".xlsx")